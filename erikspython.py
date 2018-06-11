import os

from openpyxl import load_workbook, __version__
import types
from datetime import *
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from openpyxl.styles.fonts import Font



os.chdir('/Users/ellendahlgren/Documents/')

wb = load_workbook('eriksfil2.xlsx')

sheet = wb['Data']

maxrow = sheet.max_row

wb.create_sheet(title='Data2a', index =0)
sheet = wb['Data']
sheet2a = wb['Data2a']
sheet2a.cell(row=1, column=1).value= 'Datum'
sheet2a.cell(row=1, column=2).value= 'Year'
sheet2a.cell(row=1, column=3).value= 'Month'
sheet2a.cell(row=1, column=4).value= 'Day'
sheet2a.cell(row=1, column=5).value= 'NightPM'
sheet2a.cell(row=1, column=6).value= 'NightAM'
sheet2a.cell(row=1, column=7).value= 'Dygn'
sheet2a.cell(row=1, column=8).value= 'Lying'
sheet2a.cell(row=1, column=9).value= 'Sleeping'
sheet2a.cell(row=1, column=10).value= 'Lying and Sleeping'



def _nightCheck(hour):
	if hour<= 9 or hour>=20:
		return True
	else:
		return False

rowCounter = 2

for i in range(2, maxrow+1):
    value=sheet.cell(row=i,column=1).value
    nightCheck=_nightCheck(value.hour)
    if nightCheck:
        sheet2a.cell(row=rowCounter, column=1).value = sheet.cell(row=i, column=1).value
        sheet2a.cell(row=rowCounter, column=8).value = sheet.cell(row=i, column=15).value
        sheet2a.cell(row=rowCounter, column=9).value = sheet.cell(row=i, column=16).value
        sheet2a.cell(row=rowCounter, column=2).value = value.year
        sheet2a.cell(row=rowCounter, column=3).value = value.month
        sheet2a.cell(row=rowCounter, column=4).value = value.day
        if value.hour >=19:
            sheet2a.cell(row=rowCounter, column=5).value = 1
            sheet2a.cell(row=rowCounter, column=6).value = 0
        else:
            sheet2a.cell(row=rowCounter, column=5).value = 0
            sheet2a.cell(row=rowCounter, column=6).value = 1

        lyingValue = sheet.cell(row=i, column=15).value
        sleepValue = sheet.cell(row=i, column=16).value
        if (lyingValue == 1) and (sleepValue ==1):
            sheet2a.cell(row=rowCounter, column=10).value = 1
        else:
            sheet2a.cell(row=rowCounter, column=10).value = 0

        rowCounter  = rowCounter+1


maxrow = sheet2a.max_row


dygn = 1
rowCounter = 2

for i in range(2, maxrow+1):
    value = sheet.cell(row=i, column=1).value
    hour = value.hour
    month = sheet2a.cell(row=i, column=3).value
    prevMonth = sheet2a.cell(row=i-1, column=3).value
    day = sheet2a.cell(row=i, column=4).value
    amValue = sheet2a.cell(row=i, column=6).value
    prevPmValue= sheet2a.cell(row=(i-1), column=5).value
    pmValue = sheet2a.cell(row=i, column=5).value

    if (rowCounter== 2):
        previousRow = sheet2a.cell(row=(rowCounter), column=4).value
    else:
        previousRow = sheet2a.cell(row=(rowCounter - 1), column=4).value


    if (rowCounter == 2):
        sheet2a.cell(row=rowCounter, column=7).value = dygn
    elif pmValue == 1 and day == previousRow and prevPmValue==0:
        dygn = dygn+1
        sheet2a.cell(row=rowCounter, column=7).value = dygn

    elif pmValue == 1 and day != previousRow and prevPmValue==1:
        dygn = dygn+1
        sheet2a.cell(row=rowCounter, column=7).value = dygn
    elif pmValue == 0 and prevPmValue == 0 and day != previousRow:
        dygn = dygn + 1
        sheet2a.cell(row=rowCounter, column=7).value = dygn

    elif(pmValue == prevPmValue) and (day == previousRow ):
        sheet2a.cell(row=rowCounter, column=7).value = dygn

    elif pmValue ==1 and  prevPmValue==0 and day != previousRow:
        sheet2a.cell(row=rowCounter, column=7).value = dygn

    rowCounter = rowCounter+1

wb.create_sheet(title='Data3', index =0)

sheet3 = wb['Data3']

sheet3.cell(row=1, column=1).value= 'Datum'
sheet3.column_dimensions['A'].width =25
sheet3.cell(row=1, column=2).value= 'Lying'
sheet3.cell(row=1, column=3).value= 'Sleeping'
sheet3.cell(row=1, column=4).value= 'Lying & Sleeping'
sheet3.column_dimensions['D'].width =17
sheet3.cell(row=1, column=5).value= 'Lying start time'
sheet3.column_dimensions['E'].width =25
sheet3.cell(row=1, column=6).value= 'Sleeping start time'
sheet3.column_dimensions['F'].width =25
sheet3.cell(row=1, column=7).value= 'Insomning'
sheet3.cell(row=1, column=8).value= 'Sleeping percent'
sheet3.column_dimensions['H'].width =17

for i in range(1, dygn):
    lyingCounter = 0
    sleepingCounter = 0
    lsCounter = 0
    lStartTime = 0
    sStartTime = 0


    for j in range (2, maxrow+1):
        if (sheet2a.cell(row=j, column=7).value == i) and (sheet2a.cell(row=j, column=5).value == 1):
            date = sheet2a.cell(row=j, column=1).value
        if (sheet2a.cell(row=j, column=7).value == i) and (sheet2a.cell(row=j, column=8).value == 1):
            lyingCounter = lyingCounter+1
            if (sheet2a.cell(row=(j-1), column=8).value == 0) and lStartTime == 0:
                lStartTime = sheet2a.cell(row=j, column=1).value
        if (sheet2a.cell(row=j, column=7).value == i) and (sheet2a.cell(row=j, column=9).value == 1):
            sleepingCounter = sleepingCounter + 1
            if (sheet2a.cell(row=(j-1), column=9).value == 0)and sStartTime==0:
                sStartTime = sheet2a.cell(row=j, column=1).value
        if (sheet2a.cell(row=j, column=7).value == i) and (sheet2a.cell(row=j, column=10).value == 1):
            lsCounter = lsCounter + 1
    sheet3.cell(row=i+1, column=1).value= date
    sheet3.cell(row=i+1, column=2).value= lyingCounter
    sheet3.cell(row=i+1, column=3).value= sleepingCounter
    sheet3.cell(row=i+1, column=4).value= lsCounter
    sheet3.cell(row=i+1, column=5).value = lStartTime
    sheet3.cell(row=i+1, column=6).value = sStartTime


maxrow = sheet3.max_row

for i in range (2, maxrow+1):
    sleepingPercent = 0.0
    insomning = 0.0
    sleepingPercent=0.0

    sleepingMinutes = (sheet3.cell(row=i, column=3).value*1.0)
    lyingMinutes = (sheet3.cell(row=i, column=2).value)

    if lyingMinutes !=0:
        sleepingPercent = ((sleepingMinutes/lyingMinutes)*100)
    else:
        sleepingPercent = 0

    sheet3.cell(row=i, column=8).value = sleepingPercent


from datetime import *
for i in range (2, maxrow+1):
    hour = 0
    insomning = 0

    lValue = sheet3.cell(row=i + 1, column=5).value
    sStartTime = sheet3.cell(row=i + 1, column=6).value

    if not isinstance(lValue, types.NoneType) and not isinstance(lValue, types.IntType):
        lValueHour = lValue.hour
        sStartTimeHour = sStartTime.hour

        if(lValueHour>sStartTimeHour):
            sStartTimeHour = sStartTimeHour+24


        lying = (lValueHour*60)+lValue.minute
        sleeping = (sStartTimeHour*60)+sStartTime.minute

        insomning = sleeping-lying

    sheet3.cell(row=i+1, column=7).value = insomning












wb.save('exem8.xlsx')

