import os

from openpyxl import load_workbook, __version__
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from openpyxl.styles.fonts import Font

print(__version__)

os.chdir('/Users/ellendahlgren/Documents/')
wb = load_workbook('test.xlsx')
print(wb.sheetnames)
sheet = wb['Blad1']
print(sheet['C1'].value)
sheet['C1'].value = 42
print(sheet.cell(row=1, column=3).value)

sheet['C8'].value = '=SUM(C1:C7'


sheet['B1'].font = Font(sz =14, bold=True, italic=True)


sheet.title = 'My new sheet'
print(sheet.max_row)
sheet.max_column


print(get_column_letter(25))
print(column_index_from_string('Y'))

#for i in range(1, 5):

 #   print(sheet.cell(row=i, column=3).value)

#wb.create_sheet(title='My sheet name', index =0)
sheet.row_dimensions[1].height
sheet.column_dimensions['B'].width =50
NamedStyle(name='custom_datetime', number_format='YYYY/MM/DD HH:MM:MM') sheet['A2'].style = 'custom_datetime'


#wb.save('exem2.xlsx')
#wb = openpyxl.load_workbook('test.xlsx')















